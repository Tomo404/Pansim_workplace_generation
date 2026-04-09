from __future__ import annotations

import math
import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from generator import CellKey


BIN_ORDER = ["1-4", "5-9", "10-19", "20-49", "50-249", "250-1000"]


def normalize_teaor_code(value) -> str | None:
    if pd.isna(value):
        return None

    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.lstrip("0")
    return s if s else "0"


def load_target_company_counts_by_teaor_bin(xlsx_path: Path) -> dict[tuple[str, str], int]:
    """
    Beolvassa a teaor_x_bin.xlsx KSH exportot.

    Elvárt szerkezet a screenshot alapján:
    - A oszlop: '01== ...'
    - B:G oszlopok: 1-4, 5-9, 10-19, 20-49, 50-249, 250+
    - sorok 10-től indulnak
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    out: dict[tuple[str, str], int] = {}

    col_map = {
        2: "1-4",
        3: "5-9",
        4: "10-19",
        5: "20-49",
        6: "50-249",
        7: "250-1000",
    }

    for r in range(10, ws.max_row + 1):
        raw_label = ws.cell(r, 1).value
        if raw_label is None:
            continue

        raw_label = str(raw_label).strip()
        m = re.match(r"^(\d{2})==", raw_label)
        if not m:
            continue

        teaor = normalize_teaor_code(m.group(1))

        for c, bin_name in col_map.items():
            v = ws.cell(r, c).value
            if v is None or v == "":
                v = 0

            try:
                count_val = int(round(float(v)))
            except Exception:
                count_val = 0

            out[(teaor, bin_name)] = count_val

    return out


def load_settlement_weights(xlsx_path: Path) -> pd.DataFrame:
    """
    settlement_weights.xlsx-ből:
    - settlement
    - population_share_country

    Ha a share-ek nem pont 1-re adódnak, újranormalizáljuk.
    """
    df = pd.read_excel(xlsx_path)

    required_cols = ["settlement", "population_share_country"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(
            f"A settlement_weights fájlból hiányoznak oszlopok: {missing}. "
            f"Elérhető: {list(df.columns)}"
        )

    df = df[["settlement", "population_share_country"]].copy()
    df["settlement"] = df["settlement"].astype(str).str.strip()
    df["population_share_country"] = pd.to_numeric(
        df["population_share_country"], errors="coerce"
    ).fillna(0.0)

    df = df[df["settlement"] != ""].copy()

    # ha véletlenül ugyanaz a település többször szerepelne, összevonjuk
    df = (
        df.groupby("settlement", as_index=False)["population_share_country"]
        .sum()
    )

    total_share = df["population_share_country"].sum()
    if total_share <= 0:
        raise ValueError("A settlement_weights-ben az országos share összege <= 0.")

    df["population_share_country"] = df["population_share_country"] / total_share
    return df


def aggregate_counts_by_teaor_bin(counts: dict) -> dict[tuple[str, str], int]:
    out = defaultdict(int)
    for key, value in counts.items():
        out[(str(key.teaor), str(key.bin_name))] += int(value)
    return out


def allocate_integer_counts_by_weights(
    total_to_allocate: int,
    settlements_df: pd.DataFrame,
) -> dict[str, int]:
    """
    Legnagyobb maradékos allokáció settlement szintre.
    """
    if total_to_allocate <= 0:
        return {}

    shares = settlements_df["population_share_country"].tolist()
    names = settlements_df["settlement"].tolist()

    exact = [total_to_allocate * s for s in shares]
    base = [int(math.floor(x)) for x in exact]
    remainders = [x - b for x, b in zip(exact, base)]

    missing = total_to_allocate - sum(base)

    order = sorted(range(len(names)), key=lambda i: remainders[i], reverse=True)
    for i in order[:missing]:
        base[i] += 1

    out = {}
    for name, cnt in zip(names, base):
        if cnt > 0:
            out[name] = cnt

    return out


def impute_counts_to_national_targets(
    counts: dict,
    teaor_x_bin_path: Path,
    settlement_weights_path: Path,
) -> tuple[dict, pd.DataFrame]:
    """
    A tensor counts-ot kiegészíti úgy, hogy TEÁOR × bin szinten legalább
    a teaor_x_bin.xlsx KSH országos cégszámai legyenek meg.

    A hiányzó cégeket settlement_weights alapján településekre osztja szét.
    """
    target_counts = load_target_company_counts_by_teaor_bin(teaor_x_bin_path)
    settlements_df = load_settlement_weights(settlement_weights_path)

    current_counts = aggregate_counts_by_teaor_bin(counts)
    new_counts = dict(counts)

    summary_rows = []

    all_pairs = sorted(
        set(current_counts.keys()) | set(target_counts.keys()),
        key=lambda x: (int(x[0]), BIN_ORDER.index(x[1])),
    )

    for teaor, bin_name in all_pairs:
        current_val = int(current_counts.get((teaor, bin_name), 0))
        target_val = int(target_counts.get((teaor, bin_name), 0))
        missing_val = max(0, target_val - current_val)

        if missing_val > 0:
            allocation = allocate_integer_counts_by_weights(
                total_to_allocate=missing_val,
                settlements_df=settlements_df,
            )

            for settlement, add_count in allocation.items():
                key = CellKey(
                    settlement=settlement,
                    teaor=teaor,
                    bin_name=bin_name,
                )
                new_counts[key] = int(new_counts.get(key, 0)) + int(add_count)

        summary_rows.append(
            {
                "teaor": teaor,
                "bin": bin_name,
                "current_company_count": current_val,
                "target_company_count": target_val,
                "imputed_company_count": missing_val,
            }
        )

    summary_df = pd.DataFrame(summary_rows)

    return new_counts, summary_df